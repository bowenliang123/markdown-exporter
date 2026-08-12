from openpyxl import load_workbook
from test_base import TestBase


class TestMdToXlsx(TestBase):
    def test_md_to_xlsx(self):
        # Define input and output paths
        input_file = "test/resources/example_md_table.md"
        output_file = "test_output/test.xlsx"

        # Run the tool using the base class method
        self.run_script("parser/cli_md_to_xlsx.py", input_file, output_file)

        # Verify the output file is not empty
        self.verify_output_file(output_file)

    def test_md_to_xlsx_preserves_br_line_breaks(self):
        """Regression test for GitHub issues #77 and #65:
        <br> tags inside Markdown table cells should become line breaks in XLSX cells."""
        input_file = "test/resources/example_md_table_with_br.md"
        output_file = "test_output/test_br.xlsx"

        self.run_script("parser/cli_md_to_xlsx.py", input_file, output_file)
        self.verify_output_file(output_file)

        workbook = load_workbook(output_file)
        worksheet = workbook.active
        self.assertIsNotNone(worksheet, "Worksheet is empty")
        assert worksheet is not None

        # Locate the "Description" column
        description_col: int | None = None
        for col_idx, cell in enumerate(worksheet[1], start=1):
            if cell.value == "Description":
                description_col = col_idx
                break
        self.assertIsNotNone(description_col, "Description column not found")
        assert description_col is not None

        # Cell with <br> should contain a real newline and have wrap text enabled
        item_1_description = worksheet.cell(row=2, column=description_col)
        self.assertEqual(item_1_description.value, "Line one\nLine two")
        self.assertTrue(item_1_description.alignment.wrap_text)

        # Regular cell should not contain a newline
        item_2_description = worksheet.cell(row=3, column=description_col)
        self.assertEqual(item_2_description.value, "Single line")
